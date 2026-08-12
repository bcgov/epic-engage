# Copyright © 2019 Province of British Columbia
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Tests to verify the Engagement API end-point.

Test-Suite to ensure that the /Engagement endpoint is working as expected.
"""
import copy
from datetime import datetime, timedelta
from http import HTTPStatus
from io import BytesIO
import json

from flask import current_app
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pytest

from met_api.constants.engagement_status import Status
from met_api.models.engagement import Engagement as EngagementModel
from met_api.models.membership import Membership as MembershipModel
from met_api.models.tenant import Tenant as TenantModel
from met_api.services.dashboard_export_service import (
    AGGREGATE_COLUMNS, AGGREGATE_HEADER_ROW, ALL_DATA, COMMENT_DATA_START_ROW, DASHBOARD_SHEETS,
    DATA_START_ROW, OPTION_LABEL_ROW, PAGE_TITLE_ROW, QUALITATIVE_RESPONSES,
    QUANTITATIVE_AGGREGATED, QUANTITATIVE_NON_AGGREGATED, QUESTION_TITLE_ROW, QUESTION_TYPE_ROW)
from met_api.utils.constants import TENANT_ID_HEADER
from met_api.utils.enums import ContentType, MembershipStatus
from met_api.utils.export_styles import (
    BODY_FONT_COLOUR, CHECKBOX_SELECTED_FONT_COLOUR, MUTED_FONT_COLOUR, NUMERIC_FONT_COLOUR,
    QUESTION_BANNER_COLOUR, get_page_colours, mute_colour)
from tests.utilities.factory_scenarios import (
    TestEngagementInfo, TestJwtClaims, TestParticipantInfo, TestSubmissionInfo, TestSurveyInfo,
    TestTenantInfo, TestUserInfo)
from tests.utilities.factory_utils import (
    factory_auth_header, factory_engagement_model, factory_membership_model, factory_participant_model,
    factory_staff_user_model, factory_submission_model, factory_survey_and_eng_model, factory_survey_model,
    factory_survey_report_setting_model, factory_tenant_model, set_global_tenant)


surveys_url = '/api/surveys/'


@pytest.mark.parametrize('survey_info', [TestSurveyInfo.survey1])
def test_create_survey(client, jwt, session, survey_info):  # pylint:disable=unused-argument
    """Assert that an survey can be POSTed."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    tenant_short_name = current_app.config.get('DEFAULT_TENANT_SHORT_NAME')
    headers[TENANT_ID_HEADER] = tenant_short_name
    data = {
        'name': survey_info.get('name'),
        'display': survey_info.get('form_json').get('display'),
    }
    rv = client.post(surveys_url, data=json.dumps(data),
                     headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 200
    assert rv.json.get('form_json') == survey_info.get('form_json')


def test_create_survey_with_tenant(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that an survey can be POSTed."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    tenant_short_name = current_app.config.get('DEFAULT_TENANT_SHORT_NAME')
    tenant = TenantModel.find_by_short_name(tenant_short_name)
    assert tenant is not None
    headers[TENANT_ID_HEADER] = tenant_short_name

    rv = client.post(surveys_url, data=json.dumps({
        'name': TestSurveyInfo.survey1.get('name'),
        'display': TestSurveyInfo.survey1.get('form_json').get('display'),
    }), headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 200
    survey_tenant_id = rv.json.get('tenant_id')
    assert survey_tenant_id == str(tenant.id)

    # Create a tenant
    tenant_data = TestTenantInfo.tenant2
    tenant_model = factory_tenant_model(tenant_data)
    tenant2_short_name = tenant_data['short_name']
    tenant_2 = TenantModel.find_by_short_name(tenant2_short_name)
    # Verify that the tenant was created successfully
    assert tenant_2 is not None

    # Set the tenant ID header for future requests
    headers[TENANT_ID_HEADER] = tenant2_short_name

    # Assert same staff admin can't create survey in a different tenant since he is a part of initial tenant.
    rv = client.post(surveys_url, data=json.dumps({
        'name': TestSurveyInfo.survey2.get('name'),
        'display': TestSurveyInfo.survey2.get('form_json').get('display'),
    }), headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 403

    # emulate Tenant 2 staff admin by setting tenant id
    claims = copy.deepcopy(TestJwtClaims.staff_admin_role.value)
    claims['tenant_id'] = str(tenant_model.id)
    headers = factory_auth_header(jwt=jwt, claims=claims)
    headers[TENANT_ID_HEADER] = tenant2_short_name

    # Create a survey within the new tenant
    rv = client.post(surveys_url, data=json.dumps({
        'name': TestSurveyInfo.survey2.get('name'),
        'display': TestSurveyInfo.survey2.get('form_json').get('display'),
    }), headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 200

    # Verify that the new survey belongs to the correct tenant
    survey_tenant_id = rv.json.get('tenant_id')
    assert survey_tenant_id == str(tenant_2.id)


def test_cross_tenant_edit_delete_unlinked_survey_forbidden(client, jwt, session):  # pylint:disable=unused-argument
    """Assert editing/deleting an unlinked survey outside the caller's tenant is rejected.

    `update`/`delete` on a survey with no engagement used to skip the tenant check
    entirely, so a caller whose request is not tenant-scoped (and therefore bypasses the
    tenant filter in `find_by_id`) could edit/delete a survey owned by another tenant.
    """
    # An unlinked survey owned by tenant 1.
    survey = factory_survey_model()
    survey.tenant_id = 1
    survey.save()
    survey_id = survey.id

    # A caller without a tenant claim: their by-id lookup is not tenant-scoped, so the
    # survey is resolvable, but the survey's own tenant must still gate edit/delete.
    claims = copy.deepcopy(TestJwtClaims.staff_admin_role.value)
    claims.pop('tenant_id', None)
    headers = factory_auth_header(jwt=jwt, claims=claims)

    rv = client.put(surveys_url, data=json.dumps({'id': str(survey_id), 'name': 'hijacked'}),
                    headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == HTTPStatus.FORBIDDEN

    rv = client.delete(f'{surveys_url}{survey_id}', headers=headers,
                       content_type=ContentType.JSON.value)
    assert rv.status_code == HTTPStatus.FORBIDDEN


@pytest.mark.parametrize('survey_info', [TestSurveyInfo.survey2])
def test_put_survey(client, jwt, session, survey_info):  # pylint:disable=unused-argument
    """Assert that an survey can be POSTed."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    survey = factory_survey_model()
    survey_id = str(survey.id)
    new_survey_name = 'new_survey_name'
    rv = client.put(surveys_url, data=json.dumps({'id': survey_id, 'name': new_survey_name}),
                    headers=headers, content_type=ContentType.JSON.value)

    assert rv.status_code == 200

    rv = client.get(f'{surveys_url}{survey_id}',
                    headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 200
    assert rv.json.get('form_json') == survey_info.get('form_json')
    assert rv.json.get('name') == new_survey_name


def test_survey_link(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that a survey can be POSTed."""
    survey = factory_survey_model()
    survey_id = survey.id
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)

    eng = factory_engagement_model()
    eng_id = eng.id

    # assert eng id is none in GET Survey
    rv = client.get(
        f'{surveys_url}{survey_id}',
        headers=headers,
        content_type=ContentType.JSON.value
    )
    assert rv.json.get('engagement_id') is None

    # link them together
    client.put(
        f'{surveys_url}{survey_id}/link/engagement/{eng_id}',
        headers=headers,
        content_type=ContentType.JSON.value
    )

    rv = client.get(
        f'{surveys_url}{survey_id}',
        headers=headers,
        content_type=ContentType.JSON.value
    )
    assert rv.json.get('engagement_id') == str(eng_id)


def test_get_hidden_survey_for_admins(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that a hidden survey can be fetched by admins."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    set_global_tenant()
    factory_survey_model(TestSurveyInfo.hidden_survey)

    page = 1
    page_size = 10
    sort_key = 'survey.created_date'
    sort_order = 'desc'

    rv = client.get(f'{surveys_url}?page={page}&size={page_size}&sort_key={sort_key}\
                    &sort_order={sort_order}&search_text=', headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 200
    assert rv.json.get('total') == 1


def test_get_survey_for_reviewer(client, jwt, session):  # pylint:disable=unused-argument
    """Assert reviewers different permission."""
    staff_1 = dict(TestUserInfo.user_staff_1)
    user = factory_staff_user_model(user_info=staff_1)
    claims = copy.deepcopy(TestJwtClaims.reviewer_role.value)
    claims['sub'] = str(user.external_id)
    headers = factory_auth_header(jwt=jwt, claims=claims)
    set_global_tenant()
    survey1 = factory_survey_model(TestSurveyInfo.survey1)

    # Attempt to access unlinked survey
    rv = client.get(f'{surveys_url}{survey1.id}',
                    headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 403

    # Link to a draft engagement
    eng: EngagementModel = factory_engagement_model(status=Status.Draft.value)
    survey1.engagement_id = eng.id
    survey1.commit()

    # Attempt to access survey linked to draft engagement
    rv = client.get(f'{surveys_url}{survey1.id}',
                    headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 403

    # Add user as a reviewer in the team
    factory_membership_model(user_id=user.id, engagement_id=eng.id, member_type='REVIEWER')

    # Assert Reviewer can see the survey since he is added to the team.
    rv = client.get(f'{surveys_url}{survey1.id}',
                    headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 200

    # Deactivate membership
    membership_model: MembershipModel = MembershipModel.find_by_engagement_and_user_id(eng.id, user.id)
    membership_model.status = MembershipStatus.INACTIVE.value
    membership_model.commit()

    rv = client.get(f'{surveys_url}{survey1.id}',
                    headers=headers, content_type=ContentType.JSON.value)
    # Verify reviewer lost access after being removed from the team
    assert rv.status_code == 403

    # Publish the engagement
    eng.status_id = Status.Published.value
    eng.commit()
    rv = client.get(f'{surveys_url}{survey1.id}',
                    headers=headers, content_type=ContentType.JSON.value)

    # Assert user can access  the survey even when he is removed from the team since its published.
    assert rv.status_code == 200


def test_get_hidden_survey_for_team_member(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that a hidden survey cannot be fetched by team members."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.team_member_role)
    set_global_tenant()
    factory_survey_model(TestSurveyInfo.hidden_survey)

    page = 1
    page_size = 10
    sort_key = 'survey.created_date'
    sort_order = 'desc'

    rv = client.get(f'{surveys_url}?page={page}&size={page_size}&sort_key={sort_key}\
                    &sort_order={sort_order}&search_text=', headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 200
    assert rv.json.get('total') == 0


def test_get_template_survey(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that a hidden survey cannot be fetched by team members."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    set_global_tenant()
    factory_survey_model(TestSurveyInfo.survey_template)

    page = 1
    page_size = 10
    sort_key = 'survey.created_date'
    sort_order = 'desc'

    rv = client.get(f'{surveys_url}?page={page}&size={page_size}&sort_key={sort_key}\
                    &sort_order={sort_order}&search_text=', headers=headers, content_type=ContentType.JSON.value)
    assert rv.status_code == 200
    assert rv.json.get('total') == 1


def test_edit_template_survey_for_admins(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that a hidden survey cannot be fetched by team members."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    survey = factory_survey_model(TestSurveyInfo.survey_template)
    survey_id = str(survey.id)
    new_survey_name = 'new_survey_name'
    rv = client.put(surveys_url, data=json.dumps({'id': survey_id, 'name': new_survey_name}),
                    headers=headers, content_type=ContentType.JSON.value)

    assert rv.status_code == 200, 'Admins are able to edit template surveys'


def test_edit_template_survey_for_team_member(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that a hidden survey cannot be fetched by team members."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.team_member_role)
    survey = factory_survey_model(TestSurveyInfo.survey_template)
    survey_id = str(survey.id)
    new_survey_name = 'new_survey_name'
    rv = client.put(surveys_url, data=json.dumps({'id': survey_id, 'name': new_survey_name}),
                    headers=headers, content_type=ContentType.JSON.value)

    assert rv.status_code == 403, 'Team members are not able to edit template surveys, so throws exception.'


@pytest.mark.parametrize('survey_info', [TestSurveyInfo.survey2])
def test_surveys_clone_admin(mocker, client, jwt, session, survey_info):
    """Assert that a survey can be cloned."""
    survey = factory_survey_model()
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)

    # Prepare test data
    request_data = {'name': 'New Survey'}

    # Mock the SurveyService.get method
    mocker.patch(
        'met_api.services.survey_service.SurveyService.get',
        return_value=survey_info
    )

    # Make a POST request to the SurveysClone endpoint
    response = client.post(
        f'{surveys_url}{survey.id}/clone',
        data=json.dumps(request_data),
        headers=headers,
        content_type=ContentType.JSON.value
    )

    # Assert the response status code and data
    assert response.status_code == HTTPStatus.OK
    assert response.get_json().get('form_json') == survey.form_json


@pytest.mark.parametrize('survey_info', [TestSurveyInfo.survey2])
def test_surveys_clone_team_member(mocker, client, jwt, session, survey_info):
    """Assert that a survey can be cloned."""
    survey = factory_survey_model()
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.team_member_role)

    # Prepare test data
    request_data = {'name': 'New Survey'}

    # Mock the SurveyService.get method
    mocker.patch(
        'met_api.services.survey_service.SurveyService.get',
        return_value=survey_info
    )

    # Make a POST request to the SurveysClone endpoint
    response = client.post(
        f'{surveys_url}{survey.id}/clone',
        data=json.dumps(request_data),
        headers=headers,
        content_type=ContentType.JSON.value
    )

    # Assert the response status code and data
    assert response.status_code == HTTPStatus.OK
    assert response.get_json().get('form_json') == survey.form_json


wizard_survey_info = {
    **TestSurveyInfo.survey1.value,
    'form_json': {
        'display': 'wizard',
        'components': [
            {
                'type': 'panel', 'title': 'Page 1', 'key': 'page1', 'input': False,
                'components': [{'key': 'question1', 'input': True, 'type': 'simpletextfield'}],
            },
            {
                'type': 'panel', 'title': 'Page 2', 'key': 'page2', 'input': False,
                'components': [{'key': 'question2', 'input': True, 'type': 'simplecheckboxes'}],
            },
        ],
    },
}

wizard_survey_info_with_conditional = {
    **TestSurveyInfo.survey1.value,
    'form_json': {
        'display': 'wizard',
        'components': [
            {
                'type': 'panel', 'title': 'Page 1', 'key': 'page1', 'input': False,
                'components': [
                    {
                        'key': 'question1', 'input': True, 'type': 'simpleradios',
                        'label': 'How did you hear about us?',
                        'values': [
                            {'value': 'yes', 'label': 'Yes'},
                            {'value': 'other', 'label': 'Other'},
                        ],
                    },
                    {
                        'key': 'followup1', 'input': True, 'type': 'simpletextarea',
                        'label': 'Please specify',
                        'customConditional': "show = data.question1 === 'other';",
                    },
                ],
            },
        ],
    },
}


def test_get_survey_dashboard(client, session):  # pylint:disable=unused-argument
    """Assert that the dashboard endpoint returns the wizard page structure without auth."""
    survey, _ = factory_survey_and_eng_model(wizard_survey_info)

    rv = client.get(f'{surveys_url}{survey.id}/dashboard', content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.OK
    assert rv.json.get('display') == 'wizard'
    pages = rv.json.get('pages')
    assert pages == [
        {'title': 'Page 1', 'questions': ['question1']},
        {'title': 'Page 2', 'questions': ['question2']},
    ]
    assert rv.json.get('conditional_links') == {}


def test_get_survey_dashboard_non_wizard(client, session):  # pylint:disable=unused-argument
    """Assert that a single page survey returns no pages for the dashboard."""
    survey, _ = factory_survey_and_eng_model()

    rv = client.get(f'{surveys_url}{survey.id}/dashboard', content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.OK
    assert rv.json.get('pages') == []
    assert rv.json.get('conditional_links') == {}


def test_get_survey_dashboard_conditional_links(client, session):  # pylint:disable=unused-argument
    """Assert that a conditionally-shown follow-up question is grouped under its trigger."""
    survey, _ = factory_survey_and_eng_model(wizard_survey_info_with_conditional)

    rv = client.get(f'{surveys_url}{survey.id}/dashboard', content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.OK
    assert rv.json.get('conditional_links') == {
        'followup1': {
            'trigger_key': 'question1',
            'trigger_label': 'How did you hear about us?',
            'row_key': None,
            'row_label': None,
            'trigger_values': ['other'],
            'trigger_value_labels': ['Other'],
            'follow_up_label': 'Please specify',
        },
    }


def test_get_survey_dashboard_excluded_question_has_no_conditional_link(
        client, session):  # pylint:disable=unused-argument
    """Assert that a follow-up excluded from the report is not sent to the dashboard.

    Charts and comments already drop an excluded question - the analytics result and the
    grouped-comment query both filter on the report setting - so leaving its conditional link in
    would render it as a block with nothing in it. The export still reports on it.
    """
    survey, _ = factory_survey_and_eng_model(wizard_survey_info_with_conditional)
    factory_survey_report_setting_model({
        'survey_id': survey.id,
        'question_id': 'followup1',
        'question_key': 'followup1',
        'question_type': 'simpletextarea',
        'question': 'Please specify',
        'display': False,
    })

    rv = client.get(f'{surveys_url}{survey.id}/dashboard', content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.OK
    assert rv.json.get('conditional_links') == {}


def test_get_survey_dashboard_displayed_question_keeps_its_conditional_link(
        client, session):  # pylint:disable=unused-argument
    """Assert that a report setting left switched on changes nothing."""
    survey, _ = factory_survey_and_eng_model(wizard_survey_info_with_conditional)
    factory_survey_report_setting_model({
        'survey_id': survey.id,
        'question_id': 'followup1',
        'question_key': 'followup1',
        'question_type': 'simpletextarea',
        'question': 'Please specify',
        'display': True,
    })

    rv = client.get(f'{surveys_url}{survey.id}/dashboard', content_type=ContentType.JSON.value)

    assert list(rv.json.get('conditional_links')) == ['followup1']


def test_get_survey_dashboard_draft_engagement(client, session):  # pylint:disable=unused-argument
    """Assert that the dashboard endpoint hides surveys of unpublished engagements."""
    eng = factory_engagement_model(status=Status.Draft.value)
    survey = factory_survey_model()
    survey.engagement_id = eng.id
    survey.save()

    rv = client.get(f'{surveys_url}{survey.id}/dashboard', content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.NOT_FOUND


def test_get_survey_dashboard_closed_engagement_visible(client, session):  # pylint:disable=unused-argument
    """Assert that a closed (but already-started) engagement's dashboard is still reachable."""
    eng = factory_engagement_model(status=Status.Closed.value)
    survey = factory_survey_model()
    survey.engagement_id = eng.id
    survey.save()

    rv = client.get(f'{surveys_url}{survey.id}/dashboard', content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.OK


def test_get_survey_dashboard_not_yet_started_engagement_hidden(client, session):  # pylint:disable=unused-argument
    """Assert that a published engagement whose start_date is still in the future is hidden."""
    future_eng_info = {
        **TestEngagementInfo.engagement1.value,
        'start_date': (datetime.today() + timedelta(days=1)).strftime('%Y-%m-%d'),
    }
    eng = factory_engagement_model(eng_info=future_eng_info, status=Status.Published.value)
    survey = factory_survey_model()
    survey.engagement_id = eng.id
    survey.save()

    rv = client.get(f'{surveys_url}{survey.id}/dashboard', content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.NOT_FOUND


def test_get_survey_dashboard_survey_not_found(client, session):  # pylint:disable=unused-argument
    """Assert that a nonexistent survey id returns 404."""
    rv = client.get(f'{surveys_url}999999999/dashboard', content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.NOT_FOUND


def test_get_survey_dashboard_sheet(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that the dashboard export returns a workbook with the four labelled sheets."""
    survey, _ = factory_survey_and_eng_model()
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)

    rv = client.get(f'{surveys_url}{survey.id}/dashboard/sheet', headers=headers,
                    content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.OK
    workbook = load_workbook(BytesIO(rv.data))
    assert workbook.sheetnames == [sheet.tab_name for sheet in DASHBOARD_SHEETS]
    # Excel caps tab names at 31 characters, so "All Data" is the shortened form of its title.
    assert [sheet.title for sheet in DASHBOARD_SHEETS] == [
        'Quantitative - Non-aggregated',
        'Quantitative - Aggregated',
        'All Data (Quantitative and Qualitative)',
        'Qualitative Responses',
    ]


def test_get_survey_dashboard_sheet_unauthorized(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that the dashboard export is not available without the export role."""
    survey, _ = factory_survey_and_eng_model()
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.public_user_role)

    rv = client.get(f'{surveys_url}{survey.id}/dashboard/sheet', headers=headers,
                    content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.UNAUTHORIZED


def test_get_survey_dashboard_sheet_survey_not_found(client, jwt, session):  # pylint:disable=unused-argument
    """Assert that exporting a nonexistent survey returns 404."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)

    rv = client.get(f'{surveys_url}999999999/dashboard/sheet', headers=headers,
                    content_type=ContentType.JSON.value)

    assert rv.status_code == HTTPStatus.NOT_FOUND


# A two page wizard covering every quantitative question type.
quantitative_survey_info = {
    **TestSurveyInfo.survey1.value,
    'form_json': {
        'display': 'wizard',
        'components': [
            {
                'title': 'Demographics', 'key': 'page1', 'type': 'panel', 'components': [
                    {
                        'key': 'age', 'type': 'simpleradios', 'label': 'What is your age?', 'input': True,
                        'values': [{'value': 'a1', 'label': '18-34'}, {'value': 'a2', 'label': '35-54'}],
                    },
                    {
                        'key': 'freq', 'type': 'simpleselect', 'label': 'How often?', 'input': True,
                        'values': [{'value': 'm', 'label': 'Monthly'}, {'value': 'w', 'label': 'Weekly'}],
                    },
                ],
            },
            {
                'title': 'Outreach', 'key': 'page2', 'type': 'panel', 'components': [
                    {
                        'key': 'reach', 'type': 'simplesurvey', 'label': 'Rate each method', 'input': True,
                        'questions': [{'value': 'email', 'label': 'Email'}, {'value': 'radio', 'label': 'Radio'}],
                        'values': [{'value': '1', 'label': 'Low'}, {'value': '5', 'label': 'High'}],
                    },
                    {
                        'key': 'rank', 'type': 'simpleranking', 'label': 'Rank these', 'input': True,
                        'statements': [{'id': 's1', 'label': 'Recreation'}, {'id': 's2', 'label': 'Wildlife'}],
                    },
                    {
                        'key': 'acts', 'type': 'simplecheckboxes', 'label': 'Which activities?', 'input': True,
                        'values': [{'value': 'fish', 'label': 'Fishing'}, {'value': 'hike', 'label': 'Hiking'}],
                    },
                    {
                        'key': 'notes', 'type': 'simpletextarea', 'label': 'Anything else?', 'input': True,
                    },
                ],
            },
        ],
    },
}


def _export_workbook(client, jwt, survey_id):
    """Fetch the dashboard export and return its non-aggregated worksheet."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    rv = client.get(f'{surveys_url}{survey_id}/dashboard/sheet', headers=headers,
                    content_type=ContentType.JSON.value)
    assert rv.status_code == HTTPStatus.OK
    return load_workbook(BytesIO(rv.data))[QUANTITATIVE_NON_AGGREGATED.tab_name]


def test_dashboard_sheet_header_structure(client, jwt, session):  # pylint:disable=unused-argument
    """Assert the four header rows: page banners, question titles, option labels and types."""
    survey, _ = factory_survey_and_eng_model(quantitative_survey_info)

    sheet = _export_workbook(client, jwt, survey.id)

    # Free-text excluded; radio/select take one column each, the rest one per row/option.
    assert sheet.max_column == 1 + 8
    assert [c.value for c in sheet[QUESTION_TITLE_ROW]] == [
        'Respondent ID', 'What is your age?', 'How often?', 'Rate each method', 'Rate each method',
        'Rank these', 'Rank these', 'Which activities?', 'Which activities?',
    ]
    # Names the specific row/option; empty for radio and drop-down.
    assert [c.value for c in sheet[OPTION_LABEL_ROW]] == [
        None, None, None, 'Email', 'Radio', 'Recreation', 'Wildlife', 'Fishing', 'Hiking',
    ]
    assert [c.value for c in sheet[QUESTION_TYPE_ROW]] == [
        'TYPE', 'RADIO', 'DROPDOWN', 'LIKERT MATRIX', 'LIKERT MATRIX',
        'RANK ORDER', 'RANK ORDER', 'CHECKBOX', 'CHECKBOX',
    ]
    # Each page is bannered across its own columns.
    assert sheet.cell(row=PAGE_TITLE_ROW, column=2).value == 'Page 1 - Demographics'
    assert sheet.cell(row=PAGE_TITLE_ROW, column=4).value == 'Page 2 - Outreach'
    assert {str(r) for r in sheet.merged_cells.ranges} == {'B1:C1', 'D1:I1'}


def test_dashboard_sheet_respondent_rows(client, jwt, session):  # pylint:disable=unused-argument
    """Assert one row per respondent, with each question type's value written as designed."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    participant = factory_participant_model()
    factory_submission_model(survey.id, eng.id, participant.id, {
        **TestSubmissionInfo.submission1.value,
        'submission_json': {
            'age': 'a2', 'freq': 'w',
            'reach': {'email': '5', 'radio': '1'},
            'rank': {'s1': 2, 's2': 1},
            'acts': {'fish': True, 'hike': False},
            'notes': 'excluded from this sheet',
        },
    })

    sheet = _export_workbook(client, jwt, survey.id)

    assert sheet.max_row == DATA_START_ROW
    assert [c.value for c in sheet[DATA_START_ROW]] == [
        'R-0001',
        '35-54',      # radio -> option label
        'Weekly',     # drop-down -> option label
        '5', '1',     # Likert -> scale code as stored
        2, 1,         # rank order -> position
        1, 0,         # checkbox -> selected / not selected
    ]


def test_dashboard_sheet_collapses_resubmissions(client, jwt, session):  # pylint:disable=unused-argument
    """Assert a participant's resubmission replaces their earlier row rather than adding one."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    participant = factory_participant_model()
    other = factory_participant_model(TestParticipantInfo.participant2)
    factory_submission_model(survey.id, eng.id, participant.id,
                             {**TestSubmissionInfo.submission1.value, 'submission_json': {'age': 'a1'}})
    factory_submission_model(survey.id, eng.id, participant.id,
                             {**TestSubmissionInfo.submission1.value, 'submission_json': {'age': 'a2'}})
    factory_submission_model(survey.id, eng.id, other.id,
                             {**TestSubmissionInfo.submission1.value, 'submission_json': {'age': 'a1'}})

    sheet = _export_workbook(client, jwt, survey.id)

    # The repeat participant keeps only their most recent answer.
    assert sheet.max_row == DATA_START_ROW + 1
    assert [sheet.cell(row=r, column=1).value for r in (DATA_START_ROW, DATA_START_ROW + 1)] == \
        ['R-0001', 'R-0002']
    assert sheet.cell(row=DATA_START_ROW, column=2).value == '35-54'
    assert sheet.cell(row=DATA_START_ROW + 1, column=2).value == '18-34'


def test_dashboard_sheet_unanswered_questions_stay_blank(client, jwt, session):  # pylint:disable=unused-argument
    """Assert a question the respondent never answered is blank, not a zero."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    participant = factory_participant_model()
    factory_submission_model(survey.id, eng.id, participant.id,
                             {**TestSubmissionInfo.submission1.value, 'submission_json': {'age': 'a1'}})

    sheet = _export_workbook(client, jwt, survey.id)

    assert sheet.cell(row=DATA_START_ROW, column=2).value == '18-34'
    # Checkbox especially must not report 0 for a question never shown - that would be
    # indistinguishable from a deliberate "not selected".
    assert [c.value for c in sheet[DATA_START_ROW]][2:] == [None] * 7


def _aggregated_sheet(client, jwt, survey_id):
    """Fetch the dashboard export and return its aggregated worksheet."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    rv = client.get(f'{surveys_url}{survey_id}/dashboard/sheet', headers=headers,
                    content_type=ContentType.JSON.value)
    assert rv.status_code == HTTPStatus.OK
    return load_workbook(BytesIO(rv.data))[QUANTITATIVE_AGGREGATED.tab_name]


def _seed_two_respondents(survey, eng):
    """Add two submissions covering every quantitative question type."""
    for answers in (
        {'age': 'a1', 'freq': 'm', 'reach': {'email': '1', 'radio': '5'},
         'rank': {'s1': 1, 's2': 2}, 'acts': {'fish': True, 'hike': False}},
        {'age': 'a1', 'freq': 'w', 'reach': {'email': '5'},
         'rank': {'s1': 2, 's2': 1}, 'acts': {'fish': True, 'hike': True}},
    ):
        factory_submission_model(survey.id, eng.id, factory_participant_model().id,
                                 {**TestSubmissionInfo.submission1.value, 'submission_json': answers})


def test_dashboard_aggregated_sheet_header_and_banners(client, jwt, session):  # pylint:disable=unused-argument
    """Assert the aligned header row, and that pages and questions each get their own banner."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    _seed_two_respondents(survey, eng)

    sheet = _aggregated_sheet(client, jwt, survey.id)

    # The design offset these by one; 'Question Key' is dropped so 'Question Type' sits
    # over the column holding the type.
    assert [c.value for c in sheet[AGGREGATE_HEADER_ROW]] == [
        label for label, _ in AGGREGATE_COLUMNS
    ]
    assert sheet.cell(row=AGGREGATE_HEADER_ROW, column=1).value == 'Question Type'
    assert sheet.cell(row=AGGREGATE_HEADER_ROW + 1, column=1).value == 'PAGE 1 - DEMOGRAPHICS'
    assert sheet.cell(row=AGGREGATE_HEADER_ROW + 1, column=1).alignment.horizontal == 'left'

    # Each question opens its own neutral, full-width banner.
    question_banner = sheet.cell(row=AGGREGATE_HEADER_ROW + 2, column=1)
    assert question_banner.value == 'What is your age?'
    assert question_banner.fill.fgColor.rgb[2:] == QUESTION_BANNER_COLOUR
    assert f'A2:{get_column_letter(len(AGGREGATE_COLUMNS))}2' in {
        str(r) for r in sheet.merged_cells.ranges
    }
    # Each page banner takes its own page colour.
    assert sheet.cell(row=AGGREGATE_HEADER_ROW + 1, column=1).fill.fgColor.rgb[2:] == \
        get_page_colours(0).banner
    page_two = [r for r in range(1, sheet.max_row + 1)
                if sheet.cell(row=r, column=1).value == 'PAGE 2 - OUTREACH']
    assert len(page_two) == 1
    assert sheet.cell(row=page_two[0], column=1).fill.fgColor.rgb[2:] == get_page_colours(1).banner


def test_dashboard_aggregated_sheet_tallies_by_question_type(client, jwt, session):  # pylint:disable=unused-argument
    """Assert each question type fills the columns the design assigns it."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    _seed_two_respondents(survey, eng)

    sheet = _aggregated_sheet(client, jwt, survey.id)
    rows = [[c.value for c in row] for row in sheet.iter_rows()]

    def row_for(question_type, option, extra=None):
        found = [r for r in rows if r[0] == question_type and
                 r[2] == option and (extra is None or extra in r)]
        assert len(found) == 1, f'expected one {question_type} row for {option}'
        return found[0]

    # Radio: one row per option, including one nobody picked.
    assert row_for('RADIO', '18-34')[3:5] == [2, 1.0]
    assert row_for('RADIO', '35-54')[3:5] == [0, 0.0]
    # Drop-down behaves the same.
    assert row_for('DROPDOWN', 'Monthly')[3:5] == [1, 0.5]

    # Likert: one row per statement and scale point. Only one respondent rated 'Radio',
    # so its denominator is 1.
    assert row_for('LIKERT MATRIX', 'Email', 'Low')[3:7] == [1, 0.5, '1', 'Low']
    assert row_for('LIKERT MATRIX', 'Radio', 'High')[3:7] == [1, 1.0, '5', 'High']

    # Rank order reports through its own columns, leaving count/percentage empty.
    recreation = row_for('RANK ORDER', 'Recreation', 'Ranked 1')
    assert recreation[3:5] == [None, None]
    assert recreation[7:10] == ['Ranked 1', 1, 0.5]

    # Checkbox counts respondents, not selections, so percentages can exceed 100%.
    assert row_for('CHECKBOX', 'Fishing')[3:5] == [2, 1.0]
    assert row_for('CHECKBOX', 'Hiking')[3:5] == [1, 0.5]


# Two free-text follow-ups sharing a label, each shown for a different selected option.
conditional_comment_survey_info = {
    **TestSurveyInfo.survey1.value,
    'form_json': {
        'display': 'wizard',
        'components': [
            {
                'title': 'Demographics', 'key': 'page1', 'type': 'panel', 'components': [
                    {'key': 'age', 'type': 'simpleradios', 'label': 'Age?', 'input': True,
                     'values': [{'value': 'a1', 'label': '18-34'}]},
                ],
            },
            {
                'title': 'Valued Components', 'key': 'page2', 'type': 'panel', 'components': [
                    {'key': 'vc', 'type': 'simpleradios', 'label': 'Component', 'input': True,
                     'values': [{'value': 'air', 'label': 'Air quality'},
                                {'value': 'wild', 'label': 'Wildlife'}]},
                    {'key': 'why1', 'type': 'simpletextarea', 'label': 'Why is this important?',
                     'input': True, 'customConditional': "show = data.vc === 'air'"},
                    {'key': 'why2', 'type': 'simpletextarea', 'label': 'Why is this important?',
                     'input': True, 'customConditional': "show = data.vc === 'wild'"},
                ],
            },
        ],
    },
}


def _all_data_sheet(client, jwt, survey_id):
    """Fetch the dashboard export and return its combined worksheet."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    rv = client.get(f'{surveys_url}{survey_id}/dashboard/sheet', headers=headers,
                    content_type=ContentType.JSON.value)
    assert rv.status_code == HTTPStatus.OK
    return load_workbook(BytesIO(rv.data))[ALL_DATA.tab_name]


def test_all_data_sheet_combines_quantitative_and_free_text(client, jwt, session):  # pylint:disable=unused-argument
    """Assert the combined sheet keeps every column of the quantitative sheet, plus free text."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    _seed_two_respondents(survey, eng)

    sheet = _all_data_sheet(client, jwt, survey.id)

    # The quantitative sheet's 8 question columns, plus the survey's one free-text question.
    assert sheet.max_column == 1 + 9
    assert [c.value for c in sheet[QUESTION_TITLE_ROW]] == [
        'Respondent ID', 'What is your age?', 'How often?', 'Rate each method', 'Rate each method',
        'Rank these', 'Rank these', 'Which activities?', 'Which activities?', 'Anything else?',
    ]
    # Free text sits in form order rather than being appended, and carries its own type label
    # with an empty option row, like radio and drop-down.
    assert [c.value for c in sheet[QUESTION_TYPE_ROW]][-1] == 'FREE TEXT'
    assert [c.value for c in sheet[OPTION_LABEL_ROW]][-1] is None


def test_all_data_sheet_keeps_every_respondent(client, jwt, session):  # pylint:disable=unused-argument
    """Assert respondents who wrote no free text still get a row, unlike the qualitative sheet."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    for answers in ({'age': 'a1', 'notes': 'First'}, {'age': 'a1'}):
        factory_submission_model(survey.id, eng.id, factory_participant_model().id,
                                 {**TestSubmissionInfo.submission1.value, 'submission_json': answers})

    sheet = _all_data_sheet(client, jwt, survey.id)

    assert sheet.max_row == DATA_START_ROW + 1
    assert [sheet.cell(row=r, column=1).value
            for r in (DATA_START_ROW, DATA_START_ROW + 1)] == ['R-0001', 'R-0002']
    # The second respondent is present with their quantitative answer but no comment.
    assert sheet.cell(row=DATA_START_ROW + 1, column=2).value == '18-34'
    assert sheet.cell(row=DATA_START_ROW + 1, column=10).value is None


def _qualitative_sheet(client, jwt, survey_id):
    """Fetch the dashboard export and return its qualitative worksheet."""
    headers = factory_auth_header(jwt=jwt, claims=TestJwtClaims.staff_admin_role)
    rv = client.get(f'{surveys_url}{survey_id}/dashboard/sheet', headers=headers,
                    content_type=ContentType.JSON.value)
    assert rv.status_code == HTTPStatus.OK
    return load_workbook(BytesIO(rv.data))[QUALITATIVE_RESPONSES.tab_name]


def test_dashboard_qualitative_sheet_structure(client, jwt, session):  # pylint:disable=unused-argument
    """Assert two header rows holding only free-text questions, and pages that keep their number."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    factory_submission_model(survey.id, eng.id, factory_participant_model().id,
                             {**TestSubmissionInfo.submission1.value,
                              'submission_json': {'age': 'a1', 'notes': 'Some feedback'}})

    sheet = _qualitative_sheet(client, jwt, survey.id)

    # Only the one free-text question earns a column; every quantitative one is left out.
    assert sheet.max_column == 2
    assert [c.value for c in sheet[QUESTION_TITLE_ROW]] == ['Respondent ID', 'Anything else?']
    # Page 1 holds no free text, so it is absent - but page 2 keeps its own number.
    assert sheet.cell(row=PAGE_TITLE_ROW, column=2).value == 'Page 2 - Outreach'
    assert sheet.cell(row=PAGE_TITLE_ROW, column=2).fill.fgColor.rgb[2:] == get_page_colours(1).banner
    assert sheet.cell(row=COMMENT_DATA_START_ROW, column=2).value == 'Some feedback'


def test_dashboard_qualitative_sheet_lists_only_commenters(client, jwt, session):  # pylint:disable=unused-argument
    """Assert respondents who wrote nothing are dropped, and the rest keep their own id."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    for answers in ({'notes': 'First'}, {'age': 'a1'}, {'notes': 'Third'}):
        factory_submission_model(survey.id, eng.id, factory_participant_model().id,
                                 {**TestSubmissionInfo.submission1.value, 'submission_json': answers})

    sheet = _qualitative_sheet(client, jwt, survey.id)

    # The middle respondent left no free text, so their id is skipped rather than reused.
    assert sheet.max_row == COMMENT_DATA_START_ROW + 1
    assert [sheet.cell(row=r, column=1).value
            for r in (COMMENT_DATA_START_ROW, COMMENT_DATA_START_ROW + 1)] == ['R-0001', 'R-0003']
    assert sheet.cell(row=COMMENT_DATA_START_ROW + 1, column=2).value == 'Third'


def test_dashboard_qualitative_sheet_labels_follow_ups(client, jwt, session):  # pylint:disable=unused-argument
    """Assert follow-ups sharing a label are told apart by the option that triggers them."""
    survey, eng = factory_survey_and_eng_model(conditional_comment_survey_info)
    factory_submission_model(survey.id, eng.id, factory_participant_model().id,
                             {**TestSubmissionInfo.submission1.value,
                              'submission_json': {'vc': 'air', 'why1': 'Air matters'}})

    sheet = _qualitative_sheet(client, jwt, survey.id)

    assert [c.value for c in sheet[QUESTION_TITLE_ROW]] == [
        'Respondent ID',
        'Why is this important? (Air quality)',
        'Why is this important? (Wildlife)',
    ]
    # The untriggered follow-up was never shown, so its cell is blank on a drained band.
    band = get_page_colours(1).band_light
    assert sheet.cell(row=COMMENT_DATA_START_ROW, column=3).value is None
    assert sheet.cell(row=COMMENT_DATA_START_ROW, column=3).fill.fgColor.rgb[2:] == mute_colour(band)


def test_dashboard_sheet_tones_answers_by_value(client, jwt, session):  # pylint:disable=unused-argument
    """Assert answers are toned by what they hold, and unanswered cells are muted."""
    survey, eng = factory_survey_and_eng_model(quantitative_survey_info)
    participant = factory_participant_model()
    factory_submission_model(survey.id, eng.id, participant.id, {
        **TestSubmissionInfo.submission1.value,
        # 'freq' is left out, so its column is unanswered.
        'submission_json': {
            'age': 'a1', 'reach': {'email': '5'}, 'acts': {'fish': True, 'hike': False},
        },
    })

    sheet = _export_workbook(client, jwt, survey.id)
    row = DATA_START_ROW

    def fill_of(column):
        return sheet.cell(row=row, column=column).fill.fgColor.rgb[2:]

    def font_of(column):
        return sheet.cell(row=row, column=column).font.color.rgb[2:]

    page_one_band = get_page_colours(0).band_light
    page_two_band = get_page_colours(1).band_light

    # Columns: 1 respondent id, 2 radio, 3 drop-down, 4-5 Likert, 6-7 rank, 8-9 checkbox.
    assert font_of(2) == BODY_FONT_COLOUR  # radio label
    assert font_of(4) == NUMERIC_FONT_COLOUR  # Likert scale value
    assert font_of(8) == CHECKBOX_SELECTED_FONT_COLOUR  # ticked checkbox
    # An unticked checkbox is muted but keeps its page band - it was answered, not a gap.
    assert sheet.cell(row=row, column=9).value == 0
    assert font_of(9) == MUTED_FONT_COLOUR
    assert fill_of(9) == page_two_band
    # The unanswered drop-down has no text to grey, so its band drains instead.
    assert sheet.cell(row=row, column=3).value is None
    assert fill_of(3) == mute_colour(page_one_band)
    assert fill_of(2) == page_one_band
