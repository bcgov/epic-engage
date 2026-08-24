# Copyright © 2021 Province of British Columbia
#
# Licensed under the Apache License, Version 2.0 (the 'License');
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an 'AS IS' BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Service for exporting internal dashboard survey data to a spreadsheet.

Builds the four sheets the dashboard's data export offers: the quantitative answers per
submission and aggregated, both combined with the free text, and the free text alone.

Every submission gets its own row and its own id. Resubmissions from one email share a
respondent number (R-0001-01, R-0001-02), sit together, and carry a count of how many came from
that email, so repeated submissions meant to skew the results are easy to spot.

An .xlsx rather than a literal .csv: a CSV is one flat text sheet, and cannot carry multiple
sheets, zebra striping or colour coding.
"""
from __future__ import annotations

from io import BytesIO
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from met_api.constants.report_setting_type import FormIoComponentType
from met_api.models.submission import Submission as SubmissionModel
from met_api.models.survey import Survey as SurveyModel
from met_api.utils.datetime import utc_datetime
from met_api.utils.export_styles import (
    AGGREGATE_HEADER_DESCRIBE_COLOUR, AGGREGATE_HEADER_LIKERT_COLOUR, AGGREGATE_HEADER_RANK_COLOUR,
    AGGREGATE_HEADER_TALLY_COLOUR, BODY_FONT_COLOUR, CELL_BORDER_COLOUR,
    CHECKBOX_SELECTED_FONT_COLOUR, MUTED_FONT_COLOUR, NUMERIC_FONT_COLOUR, QUESTION_BANNER_COLOUR,
    RESPONDENT_FONT_COLOUR, RESPONDENT_HEADER_COLOUR, RESPONDENT_HEADER_FONT_COLOUR,
    RESPONDENT_TYPE_COLOUR, RESPONDENT_TYPE_FONT_COLOUR, get_page_colours,
    get_question_type_colours, get_question_type_label, get_respondent_zebra_colour,
    get_zebra_colour, mute_colour)
from met_api.utils.survey_export_aggregates import build_aggregate_rows
from met_api.utils.survey_export_columns import (
    FREE_TEXT_TYPES, QUANTITATIVE_TYPES, build_export_columns, build_respondent_rows,
    filter_respondents_with_comments)


# Types whose answer is a number rather than an option label
NUMERIC_ANSWER_TYPES = {FormIoComponentType.SURVEY.value, FormIoComponentType.RANKING.value}


class DashboardSheet(NamedTuple):
    """A sheet in the dashboard export workbook."""

    tab_name: str
    title: str


QUANTITATIVE_NON_AGGREGATED = DashboardSheet('Quantitative - Non-agg', 'Quantitative - Non-aggregated')
QUANTITATIVE_AGGREGATED = DashboardSheet('Quantitative - Aggregated', 'Quantitative - Aggregated')
ALL_DATA = DashboardSheet('All Data', 'All Data (Quantitative and Qualitative)')
QUALITATIVE_RESPONSES = DashboardSheet('Qualitative Responses', 'Qualitative Responses')

DASHBOARD_SHEETS = (
    QUANTITATIVE_NON_AGGREGATED,
    QUANTITATIVE_AGGREGATED,
    ALL_DATA,
    QUALITATIVE_RESPONSES,
)

# The four header rows every data sheet opens with
PAGE_TITLE_ROW = 1
QUESTION_TITLE_ROW = 2
OPTION_LABEL_ROW = 3
QUESTION_TYPE_ROW = 4
DATA_START_ROW = 5

RESPONDENT_COLUMN = 1
SUBMISSION_COUNT_COLUMN = 2
FIRST_QUESTION_COLUMN = 3

# The two columns identifying who answered, ahead of any question.
RESPONDENT_COLUMNS = (RESPONDENT_COLUMN, SUBMISSION_COUNT_COLUMN)

RESPONDENT_COLUMN_WIDTH = 12
SUBMISSION_COUNT_COLUMN_WIDTH = 12
QUESTION_COLUMN_WIDTH = 22

# The qualitative sheet drops the option and type rows.
COMMENT_DATA_START_ROW = 3
COMMENT_COLUMN_WIDTH = 48

# The aggregated sheet's single header row, then banners and option rows beneath it.
AGGREGATE_HEADER_ROW = 1
AGGREGATE_COLUMN_WIDTH = 24
PERCENTAGE_FORMAT = '0.0%'

# Column headings in order, with the colour of the group each belongs to. The design offset
# these by one against the data; they are aligned here.
AGGREGATE_COLUMNS = (
    ('Question Type', AGGREGATE_HEADER_DESCRIBE_COLOUR),
    ('Question', AGGREGATE_HEADER_DESCRIBE_COLOUR),
    ('Answer Option', AGGREGATE_HEADER_DESCRIBE_COLOUR),
    ('Count', AGGREGATE_HEADER_TALLY_COLOUR),
    ('Percentage', AGGREGATE_HEADER_TALLY_COLOUR),
    ('Likert Matrix Point', AGGREGATE_HEADER_LIKERT_COLOUR),
    ('Likert Matrix Label', AGGREGATE_HEADER_LIKERT_COLOUR),
    ('Rank Position', AGGREGATE_HEADER_RANK_COLOUR),
    ('Rank Count', AGGREGATE_HEADER_RANK_COLOUR),
    ('Rank Percentage', AGGREGATE_HEADER_RANK_COLOUR),
)

# Indexes into AGGREGATE_COLUMNS holding a fraction rather than a plain number.
AGGREGATE_PERCENTAGE_COLUMNS = (4, 9)

_CENTERED = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center')
_CELL_EDGE = Side(style='thin', color=CELL_BORDER_COLOUR)
_THIN_BORDER = Border(left=_CELL_EDGE, right=_CELL_EDGE, top=_CELL_EDGE, bottom=_CELL_EDGE)


class DashboardExportService:  # pylint: disable=too-few-public-methods
    """Dashboard export management service."""

    @classmethod
    def export_dashboard_data_to_spread_sheet(cls, survey_id) -> tuple:
        """Build the internal dashboard export workbook for a survey.

        Returns the workbook as an in-memory byte stream along with a suggested filename.
        """
        survey = SurveyModel.find_by_id(survey_id)
        if not survey:
            raise KeyError(f'Survey with id {survey_id} not found')

        workbook = Workbook()

        workbook.remove(workbook.active)

        columns = build_export_columns(survey.form_json, QUANTITATIVE_TYPES)
        all_columns = build_export_columns(survey.form_json, QUANTITATIVE_TYPES | FREE_TEXT_TYPES)
        comment_columns = build_export_columns(survey.form_json, FREE_TEXT_TYPES)
        respondents = build_respondent_rows(SubmissionModel.get_by_survey_id(survey.id))

        for sheet in DASHBOARD_SHEETS:
            worksheet = workbook.create_sheet(title=sheet.tab_name)
            if sheet is QUANTITATIVE_NON_AGGREGATED:
                cls._build_non_aggregated_sheet(worksheet, columns, respondents)
            elif sheet is QUANTITATIVE_AGGREGATED:
                # Every submission is tallied, resubmissions included, so the totals agree with
                # the row counts on the other sheets and with the dashboard's own charts.
                cls._build_aggregated_sheet(
                    worksheet,
                    build_aggregate_rows(
                        survey.form_json, [row.submission for row in respondents]
                    ),
                )
            elif sheet is ALL_DATA:
                # Same shape as the non-aggregated sheet, with the free-text columns kept in.
                cls._build_non_aggregated_sheet(worksheet, all_columns, respondents)
            elif sheet is QUALITATIVE_RESPONSES:
                cls._build_qualitative_sheet(worksheet, comment_columns, respondents)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream, cls._build_file_name(survey)

    @classmethod
    def _build_non_aggregated_sheet(cls, worksheet, columns: list, respondents: list):
        """Write the one-row-per-respondent sheet: four header rows, then a row per respondent."""
        cls._write_respondent_header(worksheet)
        cls._write_page_banners(worksheet, columns)
        cls._write_question_headers(worksheet, columns)
        cls._write_respondent_rows(worksheet, columns, respondents)

        cls._set_respondent_column_widths(worksheet)
        for offset, column in enumerate(columns):
            letter = get_column_letter(FIRST_QUESTION_COLUMN + offset)
            worksheet.column_dimensions[letter].width = (
                COMMENT_COLUMN_WIDTH if column.component_type in FREE_TEXT_TYPES
                else QUESTION_COLUMN_WIDTH
            )
        # Keep the header and respondent id column in view while scrolling.
        worksheet.freeze_panes = worksheet.cell(row=DATA_START_ROW, column=FIRST_QUESTION_COLUMN)

    @classmethod
    def _build_aggregated_sheet(cls, worksheet, rows: list):
        """Write the one-row-per-option sheet, banner-separated by page and by question."""
        for index, (label, colour) in enumerate(AGGREGATE_COLUMNS):
            cls._style_header_cell(
                worksheet.cell(row=AGGREGATE_HEADER_ROW, column=index + 1, value=label),
                fill=colour,
                font_colour='FFFFFF',
                bold=True,
            )
            worksheet.column_dimensions[get_column_letter(index + 1)].width = AGGREGATE_COLUMN_WIDTH

        row_number = AGGREGATE_HEADER_ROW + 1
        current_page = None
        current_question = None
        striped_index = 0

        for entry in rows:
            if entry.page_index != current_page:
                current_page = entry.page_index
                current_question = None
                title = f'Page {entry.page_index + 1}'
                if entry.page_title:
                    title = f'{title} - {entry.page_title}'
                cls._write_aggregate_banner(
                    worksheet, row_number, title.upper(),
                    get_page_colours(entry.page_index).banner, 'FFFFFF',
                )
                row_number += 1
            if entry.question_label != current_question:
                current_question = entry.question_label
                cls._write_aggregate_banner(
                    worksheet, row_number, entry.question_label,
                    QUESTION_BANNER_COLOUR, BODY_FONT_COLOUR,
                )
                row_number += 1
                # Restarted per question so every block reads the same way.
                striped_index = 0

            cls._write_aggregate_row(worksheet, row_number, entry, striped_index)
            row_number += 1
            striped_index += 1

        worksheet.freeze_panes = worksheet.cell(row=AGGREGATE_HEADER_ROW + 1, column=1)

    @classmethod
    def _build_qualitative_sheet(cls, worksheet, columns: list, respondents: list):
        """Write the free-text sheet: page banners, question titles, then a row per commenter.

        Pages holding no free-text question simply do not appear, but the page numbers of those
        that do are left alone, so a banner can read "Page 5" with no Page 4 before it.
        """
        if not columns:
            return

        for row in (PAGE_TITLE_ROW, QUESTION_TITLE_ROW):
            for column in RESPONDENT_COLUMNS:
                cls._style_header_cell(
                    worksheet.cell(row=row, column=column),
                    fill=RESPONDENT_HEADER_COLOUR,
                    font_colour=RESPONDENT_HEADER_FONT_COLOUR,
                    bold=True,
                )
        worksheet.cell(row=QUESTION_TITLE_ROW, column=RESPONDENT_COLUMN, value='Respondent ID')
        worksheet.cell(row=QUESTION_TITLE_ROW, column=SUBMISSION_COUNT_COLUMN, value='Submissions')
        cls._set_respondent_column_widths(worksheet)

        cls._write_page_banners(worksheet, columns)
        for offset, column in enumerate(columns):
            index = FIRST_QUESTION_COLUMN + offset
            cls._style_header_cell(
                worksheet.cell(row=QUESTION_TITLE_ROW, column=index, value=column.question_label),
                fill=get_page_colours(column.page_index).header,
                font_colour='FFFFFF',
                bold=True,
            )
            worksheet.column_dimensions[get_column_letter(index)].width = COMMENT_COLUMN_WIDTH

        commenters = filter_respondents_with_comments(respondents, columns)
        for row_index, respondent in enumerate(commenters):
            row = COMMENT_DATA_START_ROW + row_index
            cls._write_identity_cells(worksheet, row, respondent)
            for offset, column in enumerate(columns):
                answer = column.read_answer(respondent.submission.submission_json)
                band = get_zebra_colour(column.page_index, row_index)
                cls._style_data_cell(
                    worksheet.cell(
                        row=row, column=FIRST_QUESTION_COLUMN + offset, value=answer
                    ),
                    fill=band if answer is not None else mute_colour(band),
                    font_colour=BODY_FONT_COLOUR if answer is not None else MUTED_FONT_COLOUR,
                )

        worksheet.freeze_panes = worksheet.cell(
            row=COMMENT_DATA_START_ROW, column=FIRST_QUESTION_COLUMN
        )

    @classmethod
    def _write_aggregate_banner(cls, worksheet, row: int, text: str, fill: str, font_colour: str):
        """Write a full-width banner row, left aligned so long titles stay readable."""
        cell = worksheet.cell(row=row, column=1, value=text)
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.font = Font(color=font_colour, bold=True, size=9)
        cell.border = _THIN_BORDER
        cell.alignment = _LEFT
        worksheet.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=len(AGGREGATE_COLUMNS)
        )
        # merge_cells styles only the top-left cell; fill the rest to keep a solid bar.
        for column in range(2, len(AGGREGATE_COLUMNS) + 1):
            filler = worksheet.cell(row=row, column=column)
            filler.fill = PatternFill('solid', fgColor=fill)
            filler.border = _THIN_BORDER

    @classmethod
    def _write_aggregate_row(cls, worksheet, row: int, entry, striped_index: int):
        """Write one option's tally, banded in its page's colours."""
        band = get_zebra_colour(entry.page_index, striped_index)
        values = (
            get_question_type_label(entry.component_type),
            entry.question_label,
            entry.answer_option,
            entry.count,
            entry.percentage,
            entry.likert_point,
            entry.likert_label,
            entry.rank_position,
            entry.rank_count,
            entry.rank_percentage,
        )
        for index, value in enumerate(values):
            cell = worksheet.cell(row=row, column=index + 1, value=value)
            cls._style_data_cell(
                cell,
                fill=band,
                font_colour=BODY_FONT_COLOUR if value is not None else MUTED_FONT_COLOUR,
            )
            if index in AGGREGATE_PERCENTAGE_COLUMNS:
                # A fraction underneath, so the % stays sortable.
                cell.number_format = PERCENTAGE_FORMAT

    @classmethod
    def _write_respondent_header(cls, worksheet):
        """Write the respondent columns' headers, kept neutral as they belong to no page."""
        for row in (PAGE_TITLE_ROW, QUESTION_TITLE_ROW, OPTION_LABEL_ROW):
            for column in RESPONDENT_COLUMNS:
                cls._style_header_cell(
                    worksheet.cell(row=row, column=column),
                    fill=RESPONDENT_HEADER_COLOUR,
                    font_colour=RESPONDENT_HEADER_FONT_COLOUR,
                    bold=True,
                )
        worksheet.cell(row=QUESTION_TITLE_ROW, column=RESPONDENT_COLUMN, value='Respondent ID')
        worksheet.cell(row=QUESTION_TITLE_ROW, column=SUBMISSION_COUNT_COLUMN, value='Submissions')
        # Only the id column names the type row; the count column just carries the band across.
        for column, value in zip(RESPONDENT_COLUMNS, ('TYPE', None)):
            cls._style_header_cell(
                worksheet.cell(row=QUESTION_TYPE_ROW, column=column, value=value),
                fill=RESPONDENT_TYPE_COLOUR,
                font_colour=RESPONDENT_TYPE_FONT_COLOUR,
                bold=True,
            )

    @classmethod
    def _write_page_banners(cls, worksheet, columns: list):
        """Write row 1: one merged, page-coloured banner spanning each page's columns."""
        for page_index, start, end in cls._page_spans(columns):
            colours = get_page_colours(page_index)
            title = f'Page {page_index + 1}'
            page_title = columns[start - FIRST_QUESTION_COLUMN].page_title
            if page_title:
                title = f'{title} - {page_title}'

            cls._style_header_cell(
                worksheet.cell(row=PAGE_TITLE_ROW, column=start, value=title),
                fill=colours.banner,
                font_colour='FFFFFF',
                bold=True,
            )
            if end > start:
                worksheet.merge_cells(
                    start_row=PAGE_TITLE_ROW, start_column=start, end_row=PAGE_TITLE_ROW, end_column=end
                )
                for column in range(start + 1, end + 1):
                    cls._style_header_cell(
                        worksheet.cell(row=PAGE_TITLE_ROW, column=column),
                        fill=colours.banner,
                        font_colour='FFFFFF',
                    )

    @classmethod
    def _write_question_headers(cls, worksheet, columns: list):
        """Write rows 2-4: question title, option label and question type per column."""
        for offset, column in enumerate(columns):
            index = FIRST_QUESTION_COLUMN + offset
            colours = get_page_colours(column.page_index)

            cls._style_header_cell(
                worksheet.cell(row=QUESTION_TITLE_ROW, column=index, value=column.question_label),
                fill=colours.header,
                font_colour='FFFFFF',
                bold=True,
            )
            cls._style_header_cell(
                worksheet.cell(row=OPTION_LABEL_ROW, column=index, value=column.option_label or None),
                fill=colours.header,
                font_colour='FFFFFF',
                italic=True,
            )
            type_fill, type_font = get_question_type_colours(column.component_type)
            cls._style_header_cell(
                worksheet.cell(
                    row=QUESTION_TYPE_ROW,
                    column=index,
                    value=get_question_type_label(column.component_type),
                ),
                fill=type_fill,
                font_colour=type_font,
                bold=True,
            )

    @classmethod
    def _write_identity_cells(cls, worksheet, row: int, respondent):
        """Write the respondent id and submission count, banded per email rather than per row.

        Every submission from one email shares a band, so a group of them reads as a single
        block and a reviewer can see at a glance that they came from the same person.
        """
        values = (respondent.respondent_id, respondent.submission_count)
        for column, value in zip(RESPONDENT_COLUMNS, values):
            cls._style_data_cell(
                worksheet.cell(row=row, column=column, value=value),
                fill=get_respondent_zebra_colour(respondent.group_index),
                font_colour=RESPONDENT_FONT_COLOUR,
            )

    @classmethod
    def _write_respondent_rows(cls, worksheet, columns: list, respondents: list):
        """Write one row per submission, zebra striped in their page's two band colours."""
        for row_index, respondent in enumerate(respondents):
            row = DATA_START_ROW + row_index

            cls._write_identity_cells(worksheet, row, respondent)
            for offset, column in enumerate(columns):
                answer = column.read_answer(respondent.submission.submission_json)
                band = get_zebra_colour(column.page_index, row_index)
                cls._style_data_cell(
                    worksheet.cell(
                        row=row, column=FIRST_QUESTION_COLUMN + offset, value=answer
                    ),
                    fill=band if answer is not None else mute_colour(band),
                    font_colour=cls._answer_font_colour(column, answer),
                )

    @staticmethod
    def _answer_font_colour(column, answer) -> str:
        """Pick an answer cell's text colour from what it holds."""
        if answer is None:
            return MUTED_FONT_COLOUR
        if column.component_type == FormIoComponentType.CHECKBOX.value:
            return CHECKBOX_SELECTED_FONT_COLOUR if answer else MUTED_FONT_COLOUR
        if column.component_type in NUMERIC_ANSWER_TYPES:
            return NUMERIC_FONT_COLOUR
        return BODY_FONT_COLOUR

    @staticmethod
    def _set_respondent_column_widths(worksheet):
        """Size the two identity columns, shared by every sheet that lists respondents."""
        widths = (RESPONDENT_COLUMN_WIDTH, SUBMISSION_COUNT_COLUMN_WIDTH)
        for column, width in zip(RESPONDENT_COLUMNS, widths):
            worksheet.column_dimensions[get_column_letter(column)].width = width

    @staticmethod
    def _page_spans(columns: list) -> list:
        """Group consecutive columns by page into (page_index, first_column, last_column)."""
        spans = []
        for offset, column in enumerate(columns):
            index = FIRST_QUESTION_COLUMN + offset
            if spans and spans[-1][0] == column.page_index:
                spans[-1][2] = index
            else:
                spans.append([column.page_index, index, index])
        return [tuple(span) for span in spans]

    @staticmethod
    def _style_header_cell(cell, fill: str, font_colour: str, bold: bool = False, italic: bool = False):
        """Style a cell in the header block. Every header row is centered."""
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.font = Font(color=font_colour, bold=bold, italic=italic, size=9)
        cell.border = _THIN_BORDER
        cell.alignment = _CENTERED
        return cell

    @staticmethod
    def _style_data_cell(cell, fill: str, font_colour: str):
        """Style a respondent's answer cell, left aligned so long labels stay readable."""
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.font = Font(color=font_colour, size=9)
        cell.border = _THIN_BORDER
        cell.alignment = _LEFT
        return cell

    @staticmethod
    def _build_file_name(survey: SurveyModel) -> str:
        """Build the download filename for a survey's dashboard export.

        Date only, in UTC: a time separator would be a colon, invalid in Windows filenames.
        The web client builds the same name for the file it saves.
        """
        engagement_name = survey.engagement.name if survey.engagement else survey.name
        timestamp = utc_datetime().strftime('%Y-%m-%d')
        return f'{engagement_name} - Dashboard Data - {timestamp}.xlsx'
